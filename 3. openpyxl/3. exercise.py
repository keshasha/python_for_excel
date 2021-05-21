#%% exercise 1
# 한 파일에서 각 sheet의 테이블에서 차트를 자동 생성하여 하나의 sheet를 새로 생성하여  저장.
# 각 sheet에서 테이블의 위치, 크기는 일정하지 않지만 1개만 존재
from openpyxl import load_workbook
from dateutil import parser
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
from openpyxl.chart.axis import DateAxis
#%%
wb = load_workbook('exercise1.xlsx')
wss = wb.sheetnames
ws = wb[wss[0]]

print(ws.max_row)
print(ws.max_column)

for row in ws.iter_rows(min_col=2, min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        res = int(''.join(list(filter(str.isdigit, cell.value))))
        print(res)


#%%
# BubbleChart class의 객체 생성
chart = LineChart()

# 플로팅 데이터 생성
data = Reference(ws, min_col=1, min_row=1, max_col=ws.max_column, max_row=ws.max_row)

# 차트 객체에 시리즈 데이터 추가
chart.add_data(data, titles_from_data=True)

# 차트 타이틀
chart.title = wss[0]

# # x축 타이틀
# chart.x_axis.title = ws['A1']

# # y축 타이틀
# chart.y_axis.title = ws['B1']

# 차트를 sheet에 추가
# 차트의 상단 좌측 코너를 E2에 고정(앵커)
ws.add_chart(chart, "E2")
# %%

wb.save('test.xlsx')
# %%
