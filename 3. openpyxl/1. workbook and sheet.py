# workbook, active 개념
# 여러 sheet를 포함 할 수 있는 하나의 xls 파일
# active : 작업을 하기 위해 하나의 sheet를 선택하는 개념

#%% openpyxl file create and save
from openpyxl import Workbook

wb = Workbook()                 # 새 워크북 생성
ws = wb.active                  # 현재 활성화된 sheet 가져옴
ws.title = "New Title"          # ws의 제목을 변경
wb.save("sample.xlsx")          # 변경 사항을 저장
wb.close()

#%%
from openpyxl import Workbook

wb = Workbook()
ws1 = wb.create_sheet("new sheet1")     # 새로운 sheet 생성
ws2 = wb.create_sheet("new sheet2", 2)  # 2번째에 sheet 생성

ws = wb["new sheet1"]           # dict형태로 sheet 접근 가능

target = wb.copy_worksheet(ws)  # sheet 복사


wb.save("sample.xlsx")
wb.close()