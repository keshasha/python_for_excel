#%% Basic file create, open and save
with open("sample.txt", "w") as f:
    read_data = f.read()
    print(read_data)

#%% openpyxl file create and save
from openpyxl import Workbook

wb = Workbook()                 # 새 워크북 생성
ws = wb.active                  # 현재 활성화된 sheet 가져옴
ws.title = "New Title"          # ws의 제목을 변경
wb.save("sample.xlsx")          # 변경 사항을 저장
wb.close()


#%% openpyxl file open
from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
