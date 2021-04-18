#%% Cell select and value modify
from openpyxl import Workbook

wb = Workbook()                 # 새 워크북 생성
ws = wb.active                  # 현재 활성화된 sheet 가져옴

c = ws['A4']                    # A4 를 가리킴
# c = 10                        # A4 값이 변경되는 것이 아님
c.value = 10                    # A4 를 10으로 변경
# ws['A4'] = 10                 # 동일한 결과

print(c)
print(dir(c))
print(c.value)

cell_range = ws['A1':'C2']      # A1:C2 셀을 선택
print(cell_range)

# wb.save("sample.xlsx")          # 변경 사항을 저장
# wb.close()

# %% fomular
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = "=SUM(A1, A2)"
